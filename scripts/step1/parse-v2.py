#!/usr/bin/env python3
"""엑셀 재파싱 v2 — 이번엔 전체 필드 + 서비스별 세분화.

2024 파일: 각 월(15컬럼 블록) 안의 각 서비스 컬럼별로 매출 파싱.
2025 파일: col 8 이용 서비스 사용, col 22~33 월별 매출.

출력:
  /tmp/2024_v2.json — [{customer, project, service, proj_start, proj_end, bill_start, bill_end, billing_method, notes, site_category, monthly:[{month,amount}]}, ...]
  /tmp/2025_v2.json — 동일 구조

핵심 원칙: 서비스별로 별도 row → 여러 서비스가 한 프로젝트에 있으면 여러 row 로 분해.
"""
import openpyxl as xl
import json
from datetime import datetime

BASE = '/Users/david/Desktop/클로드/004_CRM 웹개발/20260701_신규요청'

def to_date_str(v):
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, str) and len(v) >= 10: return v[:10]
    return None

def to_money(v):
    if v is None or v == '': return 0
    if isinstance(v, (int, float)):
        try: return round(float(v))
        except: return 0
    try: return round(float(str(v).replace(',', '').replace('₩', '').replace('원', '').strip()))
    except: return 0

def to_str(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

# ===== 2025 파일 =====
print("=== 2025 파일 재파싱 (v2) ===")
wb25 = xl.load_workbook(f'{BASE}/[경영관리] 2025년 매출 현황.xlsx', data_only=True)
ws25 = wb25['현장별 전체 매출']
rows_2025 = []
for r in range(9, ws25.max_row + 1):
    no = ws25.cell(r, 2).value
    if no is None: continue
    if isinstance(no, str) and (no.strip() == '복사' or no.strip() == ''): continue
    try: float(no)
    except (TypeError, ValueError): continue
    company = ws25.cell(r, 5).value
    if not company or not isinstance(company, str) or not company.strip(): continue
    project = ws25.cell(r, 6).value or company
    site_cat = to_str(ws25.cell(r, 7).value)  # 공공/민간
    service = to_str(ws25.cell(r, 8).value)
    monthly = []
    for m in range(12):
        amt = to_money(ws25.cell(r, 22 + m).value)
        if amt != 0: monthly.append({'month': m + 1, 'amount': amt})
    if not monthly: continue
    rows_2025.append({
        'row': r,
        'no': no,
        'company': company.strip(),
        'project': str(project).strip(),
        'site_category': site_cat,
        'service': service,
        'proj_start': to_date_str(ws25.cell(r, 3).value),
        'proj_end': to_date_str(ws25.cell(r, 4).value),
        'bill_start': to_date_str(ws25.cell(r, 9).value),
        'bill_end': to_date_str(ws25.cell(r, 10).value),
        'notes': to_str(ws25.cell(r, 11).value),
        'issue_date': to_date_str(ws25.cell(r, 12).value),
        'billing_method': to_str(ws25.cell(r, 13).value),
        'monthly': monthly,
    })
print(f"  파싱 행: {len(rows_2025)}")
tot25 = sum(m['amount'] for r in rows_2025 for m in r['monthly'])
print(f"  총액: {tot25:,}")
with open('/tmp/2025_v2.json', 'w') as f:
    json.dump(rows_2025, f, ensure_ascii=False)

# ===== 2024 파일 =====
print("\n=== 2024 파일 재파싱 (v2) — 서비스별 세분화 ===")
wb24 = xl.load_workbook(f'{BASE}/[경영관리] 2024년 매입매출 현황.xlsx', data_only=True)
ws24 = wb24['CaaS.Works 현장별 매출 현황']

# 각 월의 15컬럼 블록 (1-indexed): 32=1월시작, 47=2월시작, ...
# 각 블록 안에서 offset:
# 0:플랫폼, 1:견적서, 2:카메라, 3:카스웍스 캠, 4:카스웍스 탭, 5:스토리북, 6:3D,
# 7:전용앱, 8:홈페이지, 9:솔루션, 10:안전관리, 11:영상편집, 12:기타
# 13:월 이용료(총계-VAT제외, skip), 14:VAT포함(skip)
SERVICE_OFFSETS = [
    (0, '플랫폼'), (1, '견적서'), (2, 'AI CCTV'), (3, 'Wearable Cam'),
    (4, 'Mobile APP'), (5, 'Story Book'), (6, '3D'), (7, '전용앱'),
    (8, '홈페이지'), (9, '솔루션'), (10, '안전관리'), (11, '편집studio'),
    (12, '기타'),
]
MONTH_BLOCK_START = [32 + m * 15 for m in range(12)]  # 32, 47, 62, ..., 197

rows_2024 = []
for r in range(10, ws24.max_row + 1):
    no = ws24.cell(r, 2).value
    if no is None: continue
    if isinstance(no, str) and no.strip() == '': continue
    try: float(no)
    except (TypeError, ValueError): continue
    company = ws24.cell(r, 5).value
    if not company or not isinstance(company, str) or not company.strip(): continue

    company_type = to_str(ws24.cell(r, 6).value)  # 건설사/인테리어 등 — site_category 유사
    project = ws24.cell(r, 7).value or company
    proj_start = to_date_str(ws24.cell(r, 8).value)
    proj_end = to_date_str(ws24.cell(r, 9).value)
    bill_start = to_date_str(ws24.cell(r, 3).value)
    bill_end = to_date_str(ws24.cell(r, 4).value)
    billing_method = to_str(ws24.cell(r, 11).value)
    notes = to_str(ws24.cell(r, 14).value)

    # 서비스별 월별 매출 스캔
    svc_monthly = {}  # {service: [{month, amount}, ...]}
    for m in range(12):
        block_start = MONTH_BLOCK_START[m]
        for offset, service in SERVICE_OFFSETS:
            amt = to_money(ws24.cell(r, block_start + offset).value)
            if amt != 0:
                svc_monthly.setdefault(service, []).append({'month': m + 1, 'amount': amt})

    if not svc_monthly: continue

    # 서비스별로 별도 row 생성
    for service, monthly in svc_monthly.items():
        rows_2024.append({
            'row': r,
            'no': no,
            'company': company.strip(),
            'project': str(project).strip(),
            'site_category': company_type,  # 건설사/인테리어 등 (사용자 편집 가능)
            'service': service,
            'proj_start': proj_start,
            'proj_end': proj_end,
            'bill_start': bill_start,
            'bill_end': bill_end,
            'notes': notes,
            'billing_method': billing_method,
            'monthly': monthly,
        })

print(f"  파싱 (project × service) 행: {len(rows_2024)}")
tot24 = sum(m['amount'] for r in rows_2024 for m in r['monthly'])
print(f"  총액: {tot24:,}")

# 서비스별 분포
svcs = {}
for r in rows_2024:
    svcs[r['service']] = svcs.get(r['service'], 0) + 1
print(f"  서비스별 (project×service) 개수:")
for k, v in sorted(svcs.items(), key=lambda x: -x[1]):
    print(f"    {v:5d} · {k}")

with open('/tmp/2024_v2.json', 'w') as f:
    json.dump(rows_2024, f, ensure_ascii=False)

# ===== 검증 요약 =====
print(f"\n=== 요약 ===")
print(f"  2024: {len(rows_2024)}행, 합계 {tot24:,}원")
print(f"  2025: {len(rows_2025)}행, 합계 {tot25:,}원")

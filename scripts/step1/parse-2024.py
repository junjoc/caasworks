#!/usr/bin/env python3
"""2024 매출 엑셀 파싱 → /tmp/2024_parsed.json
- 시트: 'CaaS.Works 현장별 매출 현황'
- row 10~: 데이터
- 월별 이용료 (1-indexed): col 45, 60, 75, ..., 210 (30 + m*15)
- 회사 타입(col 6) 기반 service_type 매핑
"""
import openpyxl as xl
import json
from datetime import datetime

PATH = '/Users/david/Desktop/클로드/004_CRM 웹개발/20260701_신규요청/[경영관리] 2024년 매입매출 현황.xlsx'
wb = xl.load_workbook(PATH, data_only=True)
ws = wb['CaaS.Works 현장별 매출 현황']

# 월별 이용료 컬럼 (1-indexed): col 45(1월), 60(2월), 75, ..., 210(12월)
MONTH_COLS = [30 + m * 15 for m in range(1, 13)]  # 45, 60, ..., 210
print(f"월별 이용료 컬럼: {MONTH_COLS}")

def to_date_str(v):
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, str) and len(v) >= 10: return v[:10]
    return None

def to_money(v):
    if v is None or v == '': return 0
    if isinstance(v, (int, float)):
        try: return round(float(v))
        except: return 0
    try: return round(float(str(v).replace(',', '').replace('₩', '').replace('원', '').strip()))
    except: return 0

rows = []
for r in range(10, ws.max_row + 1):
    no = ws.cell(r, 2).value
    if no is None: continue
    try: float(no)
    except (TypeError, ValueError):
        if isinstance(no, str) and no.strip() in ('복사', ''): continue
        continue
    company = ws.cell(r, 5).value
    if not company or not isinstance(company, str) or not company.strip(): continue

    company_type = ws.cell(r, 6).value  # 건설사/인테리어/etc
    project = ws.cell(r, 7).value or company

    # 월별 이용료 파싱
    monthly = []
    for m_idx, col in enumerate(MONTH_COLS):
        amt = to_money(ws.cell(r, col).value)
        if amt > 0:
            monthly.append({'month': m_idx + 1, 'amount': amt})
    if not monthly: continue

    rows.append({
        'row': r,
        'no': no,
        'company': company.strip(),
        'company_type': (company_type or '').strip() if isinstance(company_type, str) else '',
        'project': project.strip() if isinstance(project, str) else str(project),
        'proj_start': to_date_str(ws.cell(r, 8).value),
        'proj_end': to_date_str(ws.cell(r, 9).value),
        'bill_start': to_date_str(ws.cell(r, 3).value),
        'bill_end': to_date_str(ws.cell(r, 4).value),
        'billing_method': (ws.cell(r, 11).value or '').strip() if isinstance(ws.cell(r, 11).value, str) else None,
        'notes': (ws.cell(r, 14).value or '').strip() if isinstance(ws.cell(r, 14).value, str) else None,
        'monthly': monthly,
    })

print(f"파싱된 데이터 행: {len(rows)}")

# 월별 합계
totals = {}
for r in rows:
    for m in r['monthly']:
        totals[m['month']] = totals.get(m['month'], 0) + m['amount']
print("\n월별 합계:")
for m in range(1, 13):
    print(f"  {m}월: {totals.get(m, 0):,}")
print(f"연간 합계: {sum(totals.values()):,}")

# 회사 타입 분포
types = {}
for r in rows:
    t = r['company_type'] or '(빈값)'
    types[t] = types.get(t, 0) + 1
print(f"\n회사 타입 분포:")
for k, v in sorted(types.items(), key=lambda x: -x[1]):
    print(f"  {v:5d} · {k}")

with open('/tmp/2024_parsed.json', 'w') as f:
    json.dump(rows, f, ensure_ascii=False, default=str)
print(f"\n→ /tmp/2024_parsed.json ({len(rows)} rows)")
